# -*- coding: utf-8 -*-
"""
從 EMS BOM 類 Excel 中，以含「Level」的列為標題中心列，
將該列的上一列、下一列與中心列同一欄位的值合併為欄名，並讀出資料列。
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl
import pandas as pd

DEFAULT_LEVEL_MARKER = "Level"
DEFAULT_HEADER_SEPARATOR = " "
DEFAULT_HEADER_SCAN_ROWS = 500
BOARD_KEYWORDS = (
    "MECHANICAL",
    "WIFI HIGH",
    "KEYPAD",
    "LED",
    "PACKAGE",
    "WIFI LOW",
    "PALLET",
    "MAIN BOARD",
    "POE",
    "NVME",
)
SELECTED_OUTPUT_COLUMNS = [
    "Model",
    # "Assembly",
    "Board",
    "Item",
    "Quantity",
    "Last BPA Currency",
    "Last BPA Price",
    "Lead Time",
    "MFG",
    "MPN",
    "Ecode",
    "M/S",
    "Main Source",
    "Time",
]


@dataclass(frozen=True)
class BomReadInfo:
    sheet: str
    level_row: int
    header_rows: tuple[int, int, int]
    data_start_row: int
    column_count: int


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S").rstrip(" 00:00:00").rstrip()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _is_separator_text(s: str) -> bool:
    """底線列（如 -----）視為無標題文字。"""
    if not s:
        return True
    return bool(re.fullmatch(r"[-_=.\s]+", s))


def find_level_header_row(
    ws: Any,
    marker: str = DEFAULT_LEVEL_MARKER,
    max_scan_rows: int = DEFAULT_HEADER_SCAN_ROWS,
) -> int | None:
    """回傳 1-based 列號：該列第一欄（A 欄）等於 marker 的列。找不到則 None。"""
    marker_norm = marker.strip()
    for r in range(1, min(max_scan_rows, ws.max_row or max_scan_rows) + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == marker_norm:
            return r
    return None


def merge_three_row_headers(
    row_above: Sequence[Any],
    row_center: Sequence[Any],
    row_below: Sequence[Any],
    sep: str = DEFAULT_HEADER_SEPARATOR,
) -> list[str]:
    n = max(len(row_above), len(row_center), len(row_below))
    def pad(row: Sequence[Any], length: int) -> list[Any]:
        lst = list(row)
        if len(lst) < length:
            lst.extend([None] * (length - len(lst)))
        return lst[:length]

    a = pad(row_above, n)
    c = pad(row_center, n)
    b = pad(row_below, n)
    headers: list[str] = []
    for i in range(n):
        pa, pc, pb = _cell_str(a[i]), _cell_str(c[i]), _cell_str(b[i])
        if _is_separator_text(pb):
            pb = ""
        if _is_separator_text(pa):
            pa = ""
        parts = [x for x in (pa, pc, pb) if x]
        seen: set[str] = set()
        deduped: list[str] = []
        for x in parts:
            if x not in seen:
                seen.add(x)
                deduped.append(x)
        name = sep.join(deduped) if deduped else f"Column_{i + 1}"
        headers.append(name)
    return headers


def uniquify_column_names(names: Iterable[str]) -> list[str]:
    from collections import Counter

    counts: Counter[str] = Counter()
    out: list[str] = []
    for raw in names:
        base = raw if raw else "Column"
        counts[base] += 1
        if counts[base] == 1:
            out.append(base)
        else:
            out.append(f"{base}_{counts[base]}")
    return out


def _normalize_level_for_compare(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _format_time_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.year}/{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.year}/{value.month}/{value.day}"

    text = str(value).strip()
    if not text:
        return None
    m = re.match(r"^\s*(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y}/{mo}/{d}"
    return text


def _resolve_column_name(columns: Sequence[str], preferred: str) -> str | None:
    """先找完全相等，找不到再找開頭相符（忽略大小寫）。"""
    preferred_lower = preferred.lower()
    for col in columns:
        if col.lower() == preferred_lower:
            return col
    for col in columns:
        if col.lower().startswith(preferred_lower):
            return col
    return None


def _find_marker_row_values(
    ws: Any,
    level_row: int,
    max_col: int,
    marker_text: str,
    right_span: int,
) -> tuple[Any, ...] | None:
    """在 Level 上方尋找 marker，回傳其右側指定數量的儲存格值。"""
    marker_norm = marker_text.strip().lower()
    scan_end_row = max(1, level_row - 1)
    for r in range(1, scan_end_row + 1):
        for c in range(1, max_col - right_span + 1):
            marker = ws.cell(row=r, column=c).value
            if marker is None:
                continue
            if str(marker).strip().lower() != marker_norm:
                continue
            return tuple(ws.cell(row=r, column=c + i).value for i in range(1, right_span + 1))
    return None


def _extract_date_above_level(ws: Any, level_row: int, max_col: int) -> str | None:
    """
    在 Level 列之上尋找儲存格文字為 DATE 的欄位，
    取其右側儲存格作為 Time 值。
    """
    values = _find_marker_row_values(
        ws, level_row=level_row, max_col=max_col, marker_text="DATE", right_span=1
    )
    if not values:
        return None
    return _format_time_value(values[0])


def _extract_assembly_info_above_level(
    ws: Any, level_row: int, max_col: int
) -> tuple[str | None, str | None]:
    """
    在 Level 列之上尋找儲存格文字為 Assembly: 的列，
    取右側兩欄資訊：
    - 第一欄 -> Assembly
    - 第二欄 -> 取第一個 '.' 前字串作為 Model
    """
    values = _find_marker_row_values(
        ws, level_row=level_row, max_col=max_col, marker_text="Assembly:", right_span=2
    )
    if not values:
        return None, None

    assembly_val = _cell_str(values[0])
    model_raw = _cell_str(values[1])
    assembly = assembly_val if assembly_val else None
    model = model_raw.split(".", 1)[0].strip() if model_raw else ""
    return assembly, (model or None)


def _read_header_rows(
    ws: Any, level_row: int, max_col: int
) -> tuple[list[Any], list[Any], list[Any]]:
    prev_r, next_r = level_row - 1, level_row + 1
    if prev_r < 1:
        raise ValueError("Level 標題列已是第 1 列，沒有「上一列」可合併。")

    row_above = [ws.cell(row=prev_r, column=j).value for j in range(1, max_col + 1)]
    row_center = [ws.cell(row=level_row, column=j).value for j in range(1, max_col + 1)]
    row_below = [ws.cell(row=next_r, column=j).value for j in range(1, max_col + 1)]

    return row_above, row_center, row_below


def _read_data_rows(ws: Any, data_start: int, max_col: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    # read_only 下 max_row 常為 1048576，勿用 range(max_row) 逐列掃描
    for row in ws.iter_rows(min_row=data_start, max_col=max_col, values_only=True):
        row_vals = list(row)
        if all(v is None for v in row_vals):
            continue
        rows.append(row_vals)
    return rows


def _last_non_empty_index(values: Sequence[Any]) -> int:
    """回傳 1-based 最後一個非空欄位索引，若全空回傳 0。"""
    for idx in range(len(values), 0, -1):
        if values[idx - 1] is not None and _cell_str(values[idx - 1]) != "":
            return idx
    return 0


def _effective_max_col(
    row_above: Sequence[Any],
    row_center: Sequence[Any],
    row_below: Sequence[Any],
    data_rows: Sequence[Sequence[Any]],
) -> int:
    """欄位範圍取表頭與資料的最大右邊界，避免漏掉無欄名但有資料的欄位。"""
    header_right = max(
        _last_non_empty_index(row_above),
        _last_non_empty_index(row_center),
        _last_non_empty_index(row_below),
    )
    data_right = 0
    for row in data_rows:
        data_right = max(data_right, _last_non_empty_index(row))
    return max(header_right, data_right)


def _build_board_column(df: pd.DataFrame) -> pd.DataFrame:
    def _normalize_board_from_description(desc_value: Any) -> str:
        desc_text = _cell_str(desc_value)
        if not desc_text:
            return ""
        desc_upper = desc_text.upper()
        for candidate in BOARD_KEYWORDS:
            if candidate in desc_upper:
                return candidate
        return desc_text

    level_col = _resolve_column_name(df.columns.tolist(), "Level")
    desc_col = _resolve_column_name(df.columns.tolist(), "Description")
    if not level_col or not desc_col:
        return df

    board_values: list[Any] = []
    current_board: Any = None
    for level_val, desc_val in zip(df[level_col].tolist(), df[desc_col].tolist()):
        if _normalize_level_for_compare(level_val) == "1":
            desc_text = _normalize_board_from_description(desc_val)
            current_board = desc_text if desc_text else None
        board_values.append(current_board)

    insert_at = df.columns.get_loc(desc_col) + 1
    df.insert(insert_at, "Board", board_values)
    return df


def _build_main_source_columns(df: pd.DataFrame) -> pd.DataFrame:
    item_col = _resolve_column_name(df.columns.tolist(), "Item")
    if not item_col:
        return df

    df["M/S"] = "M"
    df["Main Source"] = df[item_col]
    return df


def _rename_sub_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    從 2ND_SOURCE_TOTAL 起，每四欄為一組：
    1->Sub_i, 2->Sub_i_Currency, 3->Sub_i_Price, 4->刪除。
    """
    anchor_col = _resolve_column_name(df.columns.tolist(), "2ND_SOURCE_TOTAL")
    if not anchor_col:
        return df

    start = df.columns.get_loc(anchor_col)
    cols = df.columns.tolist()
    rename_map: dict[str, str] = {}
    drop_cols: list[str] = []

    group_idx = 1
    for idx in range(start, len(cols), 4):
        group = cols[idx : idx + 4]
        if len(group) >= 1:
            rename_map[group[0]] = f"Sub_{group_idx}"
        if len(group) >= 2:
            rename_map[group[1]] = f"Sub_{group_idx}_Currency"
        if len(group) >= 3:
            rename_map[group[2]] = f"Sub_{group_idx}_Price"
        if len(group) >= 4:
            drop_cols.append(group[3])
        group_idx += 1

    if rename_map:
        df = df.rename(columns=rename_map)
    if drop_cols:
        df = df.drop(columns=drop_cols)
    return df


def _apply_post_transforms(
    df: pd.DataFrame,
    time_value: str | None = None,
    assembly_value: str | None = None,
    model_value: str | None = None,
) -> pd.DataFrame:
    """集中管理轉換後衍生欄位流程，避免欄位互相覆蓋。"""
    df = _build_board_column(df)
    # 先做 Sub_* 重命名/刪除，再加 M/S 與 Main Source，避免被分組規則吃掉
    df = _rename_sub_columns(df)
    df = _build_main_source_columns(df)
    if assembly_value:
        df["Assembly"] = assembly_value
    if model_value:
        df["Model"] = model_value
    if time_value:
        df["Time"] = time_value
    return df


def _build_item_to_ecode_map(
    mapping_file: Path,
    sheet: str | int | None,
    level_marker: str,
    header_sep: str,
) -> tuple[dict[tuple[str, str], str], dict[tuple[str, str], str]]:
    """
    建立 Ecode 對照（優先 Model 維度，其次 Assembly 維度）：
    - (Model, Item) -> Ecode
    - (Assembly, Item) -> Ecode
    - 若未指定 sheet，會掃描對照檔所有工作表
    """
    sheet_names: list[str]
    if sheet is None:
        sheet_names = _list_workbook_sheet_names(mapping_file)
    elif isinstance(sheet, int):
        sheet_names = [_list_workbook_sheet_names(mapping_file)[sheet]]
    else:
        sheet_names = [sheet]

    grouped_by_model: dict[tuple[str, str], list[str]] = {}
    grouped_by_assembly: dict[tuple[str, str], list[str]] = {}
    processed_sheet_count = 0
    for sheet_name in sheet_names:
        try:
            mapping_df, _ = read_bom_with_merged_headers(
                mapping_file, sheet=sheet_name, level_marker=level_marker, header_sep=header_sep
            )
        except ValueError:
            # 對照檔可能包含摘要或說明頁，若無 Level 標題則略過。
            # 使用者明確指定 --ecode-sheet 時，應保留原錯誤行為。
            if sheet is not None:
                raise
            continue
        item_col = _resolve_column_name(mapping_df.columns.tolist(), "Item")
        ecode_col = _resolve_ecode_source_column(mapping_df.columns.tolist())
        model_col = _resolve_column_name(mapping_df.columns.tolist(), "Model")
        assembly_col = _resolve_column_name(mapping_df.columns.tolist(), "Assembly")
        if not item_col or not ecode_col:
            raise ValueError(
                f"對照檔工作表「{sheet_name}」缺少必要欄位，Item={item_col}, Customer PN={ecode_col}"
            )

        processed_sheet_count += 1
        sheet_name_key = _cell_str(sheet_name)
        model_values = mapping_df[model_col].tolist() if model_col else [""] * len(mapping_df)
        assembly_values = mapping_df[assembly_col].tolist() if assembly_col else [""] * len(mapping_df)
        for item_raw, ecode_raw, model_raw, assembly_raw in zip(
            mapping_df[item_col].tolist(),
            mapping_df[ecode_col].tolist(),
            model_values,
            assembly_values,
        ):
            item = _cell_str(item_raw)
            ecode = _cell_str(ecode_raw)
            if not item or not ecode:
                continue
            model_key = (_cell_str(model_raw), item)
            if model_key[0]:
                if model_key not in grouped_by_model:
                    grouped_by_model[model_key] = []
                if ecode not in grouped_by_model[model_key]:
                    grouped_by_model[model_key].append(ecode)

            assembly_key_text = _cell_str(assembly_raw) or sheet_name_key
            assembly_key = (assembly_key_text, item)
            if assembly_key not in grouped_by_assembly:
                grouped_by_assembly[assembly_key] = []
            if ecode not in grouped_by_assembly[assembly_key]:
                grouped_by_assembly[assembly_key].append(ecode)
    if processed_sheet_count == 0:
        raise ValueError("對照檔找不到可用工作表（需含 Level 標題列）")
    return (
        {k: "、".join(v) for k, v in grouped_by_model.items()},
        {k: "、".join(v) for k, v in grouped_by_assembly.items()},
    )


def _resolve_ecode_source_column(columns: Sequence[str]) -> str | None:
    ecode_col = _resolve_column_name(columns, "Customer PN")
    if ecode_col:
        return ecode_col
    for col in columns:
        col_key = col.lower().replace(" ", "")
        if "customer" in col_key and (
            "pn" in col_key or "partnumber" in col_key or "partno" in col_key
        ):
            return col
    return None


def _apply_ecode_mapping(
    df: pd.DataFrame,
    mapping_by_model: dict[tuple[str, str], str],
    mapping_by_assembly: dict[tuple[str, str], str],
) -> pd.DataFrame:
    item_col = _resolve_column_name(df.columns.tolist(), "Item")
    assembly_col = _resolve_column_name(df.columns.tolist(), "Assembly")
    model_col = _resolve_column_name(df.columns.tolist(), "Model")
    if not item_col:
        raise ValueError("主檔找不到 Item 欄位，無法回填 Ecode")
    if not assembly_col:
        raise ValueError("主檔找不到 Assembly 欄位，無法以 Assembly + Item 回填 Ecode")
    if not model_col:
        raise ValueError("主檔找不到 Model 欄位，無法以 Model + Item 回填 Ecode")

    # 同料號跨多表時，先用 Assembly 區分；僅在未命中時才回退到 Model。
    mapped = pd.Series(
        (
            (
                mapping_by_assembly.get((_cell_str(assembly_val), _cell_str(item_val)), "")
                or mapping_by_model.get((_cell_str(model_val), _cell_str(item_val)), "")
            )
            for model_val, assembly_val, item_val in zip(
                df[model_col], df[assembly_col], df[item_col]
            )
        ),
        index=df.index,
    )
    if "Ecode" in df.columns:
        old = df["Ecode"].fillna("").astype(str).str.strip()
        new = mapped.fillna("").astype(str).str.strip()
        df["Ecode"] = new.where(new != "", old)
    else:
        insert_at = df.columns.get_loc(item_col) + 1
        df.insert(insert_at, "Ecode", mapped)
    return df


def _expand_sub_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    依每列 Sub_n 展開子料列：
    - 主列保留
    - 每個 Sub_n 產生一列 Item=Sub_n, M/S='S'
    - 子列沿用主列的 Ecode/Model/Assembly/Board/Quantity/Main Source/Time/主料
    - 子列 Last BPA Currency/Price 由 Sub_n_Currency/Sub_n_Price 帶入
    """
    item_col = _resolve_column_name(df.columns.tolist(), "Item")
    ms_col = _resolve_column_name(df.columns.tolist(), "M/S")
    if not item_col or not ms_col:
        return df

    if "主料" not in df.columns:
        insert_at = df.columns.get_loc(item_col) + 1
        df.insert(insert_at, "主料", df[item_col])
    else:
        df["主料"] = df["主料"].where(df["主料"].notna(), df[item_col])

    sub_item_cols = [
        c for c in df.columns if re.fullmatch(r"Sub_\d+", c or "")
    ]
    sub_item_cols.sort(key=lambda c: int(c.split("_")[1]))
    if not sub_item_cols:
        return df

    last_bpa_currency_col = _resolve_column_name(df.columns.tolist(), "Last BPA Currency")
    last_bpa_price_col = _resolve_column_name(df.columns.tolist(), "Last BPA Price")

    expanded_rows: list[dict[str, Any]] = []
    all_cols = df.columns.tolist()
    assembly_col = _resolve_column_name(df.columns.tolist(), "Assembly")
    board_col = _resolve_column_name(df.columns.tolist(), "Board")
    model_col = _resolve_column_name(df.columns.tolist(), "Model")
    ecode_col = _resolve_column_name(df.columns.tolist(), "Ecode")
    quantity_col = _resolve_column_name(df.columns.tolist(), "Quantity")
    main_source_col = _resolve_column_name(df.columns.tolist(), "Main Source")
    time_col = _resolve_column_name(df.columns.tolist(), "Time")
    for _, row in df.iterrows():
        main_row = row.to_dict()
        main_item = _cell_str(main_row.get(item_col))
        if main_item and not _cell_str(main_row.get("主料")):
            main_row["主料"] = main_item
        expanded_rows.append(main_row)

        for sub_col in sub_item_cols:
            sub_item = _cell_str(main_row.get(sub_col))
            if not sub_item:
                continue

            # S 列只保留指定主料欄位 + 子料自身需要的欄位，其他一律清空。
            sub_row = {c: "" for c in all_cols}
            sub_row[item_col] = sub_item
            sub_row[ms_col] = "S"
            sub_row["主料"] = main_item
            if main_source_col:
                # Main Source = 主料 Item（也就是你要求的主料）
                sub_row[main_source_col] = main_item

            if ecode_col:
                sub_row[ecode_col] = main_row.get(ecode_col, "")
            if model_col:
                sub_row[model_col] = main_row.get(model_col, "")
            if assembly_col:
                sub_row[assembly_col] = main_row.get(assembly_col, "")
            if board_col:
                sub_row[board_col] = main_row.get(board_col, "")
            if quantity_col:
                sub_row[quantity_col] = main_row.get(quantity_col, "")
            if time_col:
                sub_row[time_col] = main_row.get(time_col, "")

            # 把 Last BPA Currency/Price 指到目前 Sub_n 的幣別/價格
            sub_currency_col = f"{sub_col}_Currency"
            sub_price_col = f"{sub_col}_Price"
            if last_bpa_currency_col and sub_currency_col in df.columns:
                sub_row[last_bpa_currency_col] = main_row.get(sub_currency_col, "")
            if last_bpa_price_col and sub_price_col in df.columns:
                sub_row[last_bpa_price_col] = main_row.get(sub_price_col, "")

            expanded_rows.append(sub_row)

    expanded_df = pd.DataFrame(expanded_rows, columns=all_cols)
    return expanded_df


def _update_info_column_count(info: BomReadInfo, column_count: int) -> BomReadInfo:
    return BomReadInfo(
        sheet=info.sheet,
        level_row=info.level_row,
        header_rows=info.header_rows,
        data_start_row=info.data_start_row,
        column_count=column_count,
    )


def _print_run_summary(info: BomReadInfo, row_count: int) -> None:
    print(f"工作表: {info.sheet}")
    print(f"標題列（上/中/下）: {info.header_rows}，資料自第 {info.data_start_row} 列")
    print(f"欄數: {info.column_count}，資料列數: {row_count}")


def _list_workbook_sheet_names(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets]
    finally:
        wb.close()


def read_bom_with_merged_headers(
    path: str | Path,
    sheet: str | int | None = None,
    level_marker: str = DEFAULT_LEVEL_MARKER,
    header_sep: str = DEFAULT_HEADER_SEPARATOR,
) -> tuple[pd.DataFrame, BomReadInfo]:
    """
    回傳 (DataFrame, info)。
    info 含 level_row（1-based）、columns 合併後欄名、資料起始列。
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
            used_sheet = ws.title
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
            used_sheet = ws.title
        else:
            ws = wb[sheet]
            used_sheet = sheet

        level_row = find_level_header_row(ws, marker=level_marker)
        if level_row is None:
            raise ValueError(
                f"找不到第一欄為「{level_marker}」的標題列（已掃描前 {DEFAULT_HEADER_SCAN_ROWS} 列）。"
            )

        max_col = ws.max_column or 1
        row_above, row_center, row_below = _read_header_rows(ws, level_row, max_col)
        prev_r, next_r = level_row - 1, level_row + 1
        time_value = _extract_date_above_level(ws, level_row=level_row, max_col=max_col)
        assembly_value, model_value = _extract_assembly_info_above_level(
            ws, level_row=level_row, max_col=max_col
        )

        data_start = level_row + 2
        rows = _read_data_rows(ws, data_start=data_start, max_col=max_col)
        max_col = _effective_max_col(row_above, row_center, row_below, rows)

        row_above = row_above[:max_col]
        row_center = row_center[:max_col]
        row_below = row_below[:max_col]
        rows = [row[:max_col] for row in rows]

        raw_headers = merge_three_row_headers(row_above, row_center, row_below, sep=header_sep)
        columns = uniquify_column_names(raw_headers)

        df = pd.DataFrame(rows, columns=columns)
        df = _apply_post_transforms(
            df,
            time_value=time_value,
            assembly_value=assembly_value,
            model_value=model_value,
        )

        info = BomReadInfo(
            sheet=used_sheet,
            level_row=level_row,
            header_rows=(prev_r, level_row, next_r),
            data_start_row=data_start,
            column_count=len(df.columns),
        )
        return df, info
    finally:
        wb.close()


def read_bom_multi_sheet(
    path: str | Path,
    sheet: str | int | None = None,
    level_marker: str = DEFAULT_LEVEL_MARKER,
    header_sep: str = DEFAULT_HEADER_SEPARATOR,
) -> tuple[pd.DataFrame, BomReadInfo, list[str]]:
    """
    讀取主檔：
    - 指定 sheet：單工作表模式
    - 未指定 sheet：掃描所有工作表並合併（略過不含 Level 標題的頁）
    """
    path = Path(path)
    if sheet is not None:
        df, info = read_bom_with_merged_headers(
            path, sheet=sheet, level_marker=level_marker, header_sep=header_sep
        )
        return df, info, [info.sheet]

    dfs: list[pd.DataFrame] = []
    infos: list[BomReadInfo] = []
    processed_sheets: list[str] = []
    for sheet_name in _list_workbook_sheet_names(path):
        try:
            df_sheet, info_sheet = read_bom_with_merged_headers(
                path, sheet=sheet_name, level_marker=level_marker, header_sep=header_sep
            )
        except ValueError:
            continue
        dfs.append(df_sheet)
        infos.append(info_sheet)
        processed_sheets.append(sheet_name)

    if not dfs:
        raise ValueError("主檔找不到可用工作表（需含 Level 標題列）")

    merged_df = pd.concat(dfs, ignore_index=True, sort=False)
    first_info = infos[0]
    merged_info = BomReadInfo(
        sheet=f"ALL({len(processed_sheets)})",
        level_row=first_info.level_row,
        header_rows=first_info.header_rows,
        data_start_row=first_info.data_start_row,
        column_count=len(merged_df.columns),
    )
    return merged_df, merged_info, processed_sheets


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="以含 Level 的列為中心，合併上/下一列為欄名並匯出 BOM 資料表。"
    )
    parser.add_argument("input", type=Path, help="輸入 .xlsx 路徑")
    parser.add_argument("-o", "--output", type=Path, help="輸出 .csv 或 .xlsx（依副檔名）")
    parser.add_argument("--sheet", help="工作表名稱（未指定時掃描並合併所有可用工作表）")
    parser.add_argument(
        "--sub-source",
        "--ecode-source",
        dest="sub_source",
        type=Path,
        required=False,
        help="子料來源檔（預設解讀為 EMS BOM COST；提供 Sub_* 的幣別/價格；未提供則使用 BOM COST 內建 Sub_*）",
    )
    parser.add_argument(
        "--sub-sheet",
        "--ecode-sheet",
        dest="sub_sheet",
        help="子料來源檔工作表（預設作用中工作表；未指定時掃描所有工作表）",
    )
    parser.add_argument(
        "--marker",
        default=DEFAULT_LEVEL_MARKER,
        help="用於定位標題中心列的第一欄文字（預設 Level）",
    )
    parser.add_argument(
        "--sep",
        default=DEFAULT_HEADER_SEPARATOR,
        help="合併三列標題時的分隔字串（預設空白字元）",
    )
    parser.add_argument(
        "--preview",
        action="store_true",
        help="僅印出合併後欄名與列資訊，不寫檔",
    )
    output_mode_group = parser.add_mutually_exclusive_group()
    output_mode_group.add_argument(
        "--selected-only",
        action="store_true",
        help="只輸出指定欄位檔（*_selected）",
    )
    output_mode_group.add_argument(
        "--no-selected",
        action="store_true",
        help="只輸出完整欄位檔（不輸出 *_selected）",
    )
    return parser


def _print_preview(df: pd.DataFrame, info: BomReadInfo) -> None:
    _print_run_summary(info, len(df))
    print("\n合併欄名（前 15 欄）:")
    for i, col in enumerate(df.columns[:15], 1):
        print(f"  {i}. {col}")
    print("\n前 3 列預覽:")
    print(df.head(3).to_string())


def _write_output(df: pd.DataFrame, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    suffix = output.suffix.lower()
    if suffix == ".csv":
        df.to_csv(output, index=False, encoding="utf-8-sig")
        return
    if suffix in (".xlsx", ".xlsm"):
        df.to_excel(output, index=False, engine="openpyxl")
        return
    raise ValueError("輸出副檔名請使用 .csv 或 .xlsx")


def _build_selected_columns_df(
    df: pd.DataFrame, selected_columns: Sequence[str]
) -> pd.DataFrame:
    """
    依指定欄位順序建立輸出 DataFrame。
    欄位不存在時補空字串，確保輸出欄位固定。
    """
    out = pd.DataFrame(index=df.index)
    for col in selected_columns:
        out[col] = df[col] if col in df.columns else ""
    return out


def _build_selected_output_path(output: Path) -> Path:
    return output.with_name(f"{output.stem}_selected{output.suffix}")


def _read_main_dataframe(args: argparse.Namespace) -> tuple[pd.DataFrame, BomReadInfo]:
    """讀取主檔（單表或多表），並回傳合併後資料與摘要資訊。"""
    sheet_arg: str | int | None = args.sheet if args.sheet else None
    df, info, processed_main_sheets = read_bom_multi_sheet(
        args.input,
        sheet=sheet_arg,
        level_marker=args.marker,
        header_sep=args.sep,
    )
    print(f"主檔工作表數: {len(processed_main_sheets)}")
    return df, info


def _apply_ecode_if_needed(
    df: pd.DataFrame, info: BomReadInfo, args: argparse.Namespace
) -> tuple[pd.DataFrame, BomReadInfo]:
    """若提供對照檔則回填 Ecode；未提供時直接回傳原資料。"""
    if not args.ecode_source:
        return df, info
    if not args.ecode_source.is_file():
        raise ValueError(f"找不到 Ecode 對照檔: {args.ecode_source}")

    ecode_sheet_arg: str | int | None = args.ecode_sheet if args.ecode_sheet else None
    ecode_map_by_model, ecode_map_by_assembly = _build_item_to_ecode_map(
        mapping_file=args.ecode_source,
        sheet=ecode_sheet_arg,
        level_marker=args.marker,
        header_sep=args.sep,
    )
    df = _apply_ecode_mapping(df, ecode_map_by_model, ecode_map_by_assembly)
    print(
        "Ecode 對照筆數: "
        f"Model+Item={len(ecode_map_by_model)}, "
        f"Assembly+Item={len(ecode_map_by_assembly)}"
    )
    return df, _update_info_column_count(info, len(df.columns))


def _apply_final_transforms(df: pd.DataFrame, info: BomReadInfo) -> tuple[pd.DataFrame, BomReadInfo]:
    """集中最終資料整理：展開 Sub，並移除 Item 空白列。"""
    # BOM COST 模式下必須存在 Sub_*，否則無法產生 S 子料列。
    sub_item_cols = [c for c in df.columns if re.fullmatch(r"Sub_\d+", str(c or ""))]
    if not sub_item_cols:
        raise ValueError("BOM COST 找不到 Sub_* 欄位（確認 2ND_SOURCE_TOTAL 起的分組欄位已被正確 rename）。")
    df = _expand_sub_rows(df)
    df = _drop_empty_item_rows(df)
    return df, _update_info_column_count(info, len(df.columns))


def _drop_empty_item_rows(df: pd.DataFrame) -> pd.DataFrame:
    """刪除 Item 為空/空白字元的資料列，避免輸出無效項目。"""
    item_col = _resolve_column_name(df.columns.tolist(), "Item")
    if not item_col:
        return df
    mask = df[item_col].fillna("").astype(str).str.strip() != ""
    return df.loc[mask].reset_index(drop=True)


def _ensure_ecode_column_from_bom(df: pd.DataFrame) -> pd.DataFrame:
    """
    BOM COST 本身含 Ecode（通常是 Customer PN 欄）。
    - 若已存在 Ecode：不處理
    - 否則嘗試把 Customer PN 欄改名為 Ecode
    """
    if "Ecode" in df.columns:
        return df
    customer_pn_col = _resolve_column_name(df.columns.tolist(), "Customer PN")
    if customer_pn_col:
        return df.rename(columns={customer_pn_col: "Ecode"})
    return df


def _ensure_unique_bom_ems_key(
    df: pd.DataFrame,
    df_name: str,
    key_cols: tuple[str, str, str],
) -> None:
    """
    確保 (Assembly, Model, Item) 對應唯一，若重複就報錯（依需求：警告並要求唯一）。
    """
    a_col, m_col, i_col = key_cols
    if not all(c in df.columns for c in key_cols):
        raise ValueError(f"{df_name} 找不到必要 key 欄位: {key_cols}")

    key_df = df[list(key_cols)].fillna("").astype(str)
    key_series = (
        key_df[a_col].str.strip()
        + "||"
        + key_df[m_col].str.strip()
        + "||"
        + key_df[i_col].str.strip()
    )
    dup_mask = key_series.duplicated(keep=False)
    if dup_mask.any():
        # 顯示前幾個重複 key，便於使用者追查
        dup_keys = key_series[dup_mask].unique().tolist()[:5]
        raise ValueError(f"{df_name} 在 (Assembly, Model, Item) 出現重複 key，例子: {dup_keys}")


def _expand_sub_rows_from_ems_using_bom_template(
    df_bom: pd.DataFrame,
    df_ems: pd.DataFrame,
) -> pd.DataFrame:
    """
    產生最終扁平化資料：
    - 主料列 (M)：直接使用 BOM COST 主料（M/S 已為 M）
    - 子料列 (S)：依 EMS BOM COST 的 Sub_n 展開；幣別/價格取 Sub_n_Currency/Sub_n_Price
    - 其他欄位沿用 BOM COST 主料列（以 Assembly+Model+Item 對應）
    - 同一 key 重複會直接報錯（依需求）
    """
    item_col = _resolve_column_name(df_bom.columns.tolist(), "Item")
    ms_col = _resolve_column_name(df_bom.columns.tolist(), "M/S")
    model_col = _resolve_column_name(df_bom.columns.tolist(), "Model")
    board_col = _resolve_column_name(df_bom.columns.tolist(), "Board")
    time_col = _resolve_column_name(df_bom.columns.tolist(), "Time")
    assembly_col = _resolve_column_name(df_bom.columns.tolist(), "Assembly")
    quantity_col = _resolve_column_name(df_bom.columns.tolist(), "Quantity")
    main_source_col = _resolve_column_name(df_bom.columns.tolist(), "Main Source")
    ecode_col = _resolve_column_name(df_bom.columns.tolist(), "Ecode")

    if not all([item_col, ms_col, model_col, board_col]):
        raise ValueError(
            f"BOM COST 缺少必要欄位（Item/M/S/Board/Model），目前: Item={item_col}, M/S={ms_col}, Board={board_col}, Model={model_col}"
        )

    last_cur_col = _resolve_column_name(df_bom.columns.tolist(), "Last BPA Currency")
    last_price_col = _resolve_column_name(df_bom.columns.tolist(), "Last BPA Price")
    if not last_cur_col or not last_price_col:
        raise ValueError(
            f"BOM COST 缺少 Last BPA Currency/Price 欄位，Last BPA Currency={last_cur_col}, Last BPA Price={last_price_col}"
        )

    # 基於相同規則也要求 EMS 具備必要 key 與 Sub_n 欄位
    sub_item_cols = [c for c in df_ems.columns if re.fullmatch(r"Sub_\d+", str(c or ""))]
    sub_item_cols.sort(key=lambda c: int(str(c).split("_")[1]))
    if not sub_item_cols:
        raise ValueError("EMS BOM COST 找不到 Sub_* 欄位（已確認 2ND_SOURCE_TOTAL 起的分組應已被 rename）")

    m2_col = _resolve_column_name(df_ems.columns.tolist(), "Model") or model_col
    i2_col = _resolve_column_name(df_ems.columns.tolist(), "Item") or item_col
    board2_col = _resolve_column_name(df_ems.columns.tolist(), "Board") or board_col

    _ensure_unique_bom_ems_key(df_bom, "BOM COST", (board_col, model_col, item_col))
    _ensure_unique_bom_ems_key(df_ems, "EMS BOM COST", (board2_col, m2_col, i2_col))

    # 準備 key -> row lookup（EMS 用）
    ems_key_to_row: dict[tuple[str, str, str], dict[str, Any]] = {}
    for _, r in df_ems.iterrows():
        key = (
            _cell_str(
                r.get(board_col) if board_col in df_ems.columns else r.get(board2_col)
            ),
            _cell_str(r.get(model_col) if model_col in df_ems.columns else r.get(m2_col)),
            _cell_str(r.get(item_col) if item_col in df_ems.columns else r.get(i2_col)),
        )
        ems_key_to_row[key] = r.to_dict()

    # 主料列 + 子料列
    all_cols = df_bom.columns.tolist()
    # 保持與原本 expand_sub_rows 一致：新增 主料 欄
    if "主料" not in df_bom.columns:
        insert_at = df_bom.columns.get_loc(item_col) + 1
        df_bom = df_bom.copy()
        df_bom.insert(insert_at, "主料", df_bom[item_col])
        all_cols = df_bom.columns.tolist()

    rows_out: list[dict[str, Any]] = []

    for _, bom_row in df_bom.iterrows():
        bom_item = _cell_str(bom_row.get(item_col))
        key = (_cell_str(bom_row.get(board_col)), _cell_str(bom_row.get(model_col)), bom_item)

        # M row：保留 BOM 主料原始內容
        main_dict = bom_row.to_dict()
        main_dict["主料"] = bom_item
        rows_out.append(main_dict)

        # S rows：從 EMS 的同 key 展開
        ems_row = ems_key_to_row.get(key)
        if ems_row is None:
            continue

        for sub_item_col in sub_item_cols:
            sub_item = _cell_str(ems_row.get(sub_item_col))
            if not sub_item:
                continue

            sub_currency_col = f"{sub_item_col}_Currency"
            sub_price_col = f"{sub_item_col}_Price"
            sub_currency = ems_row.get(sub_currency_col, "")
            sub_price = ems_row.get(sub_price_col, "")

            # S row：只保留你指定的主料欄位；其餘全部留空。
            child_dict = {c: "" for c in all_cols}
            child_dict[item_col] = sub_item
            child_dict[ms_col] = "S"
            child_dict[last_cur_col] = sub_currency if sub_currency is not None else ""
            child_dict[last_price_col] = sub_price if sub_price is not None else ""
            child_dict["主料"] = bom_item

            # Main Source = 主料 Item
            if main_source_col:
                child_dict[main_source_col] = bom_item

            if ecode_col:
                child_dict[ecode_col] = bom_row.get(ecode_col, "")
            if model_col:
                child_dict[model_col] = bom_row.get(model_col, "")
            if assembly_col:
                child_dict[assembly_col] = bom_row.get(assembly_col, "")
            if board_col:
                child_dict[board_col] = bom_row.get(board_col, "")
            if quantity_col:
                child_dict[quantity_col] = bom_row.get(quantity_col, "")
            if time_col:
                child_dict[time_col] = bom_row.get(time_col, "")

            rows_out.append(child_dict)

    out_df = pd.DataFrame(rows_out, columns=all_cols)
    return out_df


def _write_outputs(df: pd.DataFrame, args: argparse.Namespace) -> None:
    out = args.output or args.input.with_name(args.input.stem + "_flat.csv")
    selected_out = _build_selected_output_path(out)
    selected_df = _build_selected_columns_df(df, SELECTED_OUTPUT_COLUMNS)
    write_full_output = not args.selected_only
    write_selected_output = not args.no_selected

    if write_full_output:
        _write_output(df, out)
    if write_selected_output:
        _write_output(selected_df, selected_out)

    if write_full_output:
        if args.output:
            print(f"已寫入: {out.resolve()}")
        else:
            print(f"未指定 -o，已寫入: {out.resolve()}")
    if write_selected_output:
        if args.output:
            print(f"已寫入指定欄位檔: {selected_out.resolve()}")
        else:
            print(f"未指定 -o，已寫入指定欄位檔: {selected_out.resolve()}")


def main(argv: list[str] | None = None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)

    if not args.input.is_file():
        print(f"找不到檔案: {args.input}", file=sys.stderr)
        return 1

    try:
        df, info = _read_main_dataframe(args)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    try:
        # BOM COST 當作主料來源：Ecode 直接取自 BOM COST（不做 mapping）
        df = _ensure_ecode_column_from_bom(df)
        if "Ecode" not in df.columns:
            raise ValueError("BOM COST 缺少 Ecode 欄位，請確認是否有 Ecode 或 Customer PN 欄可供轉成 Ecode")

        # 先刪除 Item 空白列，避免後續 key 對應爆掉
        df = _drop_empty_item_rows(df)

        if args.sub_source:
            # （進階）若提供 EMS 子料來源，則用 EMS 的 Sub_* 生成 S 子料列
            sub_sheet_arg: str | int | None = args.sub_sheet if args.sub_sheet else None
            df_ems, _, processed_sub_sheets = read_bom_multi_sheet(
                args.sub_source,
                sheet=sub_sheet_arg,
                level_marker=args.marker,
                header_sep=args.sep,
            )
            print(f"子料來源工作表數: {len(processed_sub_sheets)}")
            df_ems = _drop_empty_item_rows(df_ems)

            df = _expand_sub_rows_from_ems_using_bom_template(df, df_ems)
            info = _update_info_column_count(info, len(df.columns))
        else:
            # 預設：只使用 BOM COST 內建 Sub_* 展開
            df, info = _apply_final_transforms(df, info)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    if args.preview:
        _print_preview(df, info)
        return 0
    _print_run_summary(info, len(df))

    try:
        _write_outputs(df, args)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
